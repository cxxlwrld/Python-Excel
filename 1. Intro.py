import openpyxl
import itertools
# create a new workbook
workbook = openpyxl.Workbook()

# select the active worksheet
ws = workbook.active

# Add data to the worksheet
ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['A2'] = 'John'
ws['B2'] = 25
ws['A3'] = 'Mary'
ws['B3'] = 30

# save the workbook
workbook.save('example.csv')

# read data from the worksheet
ws = workbook.active
for row in ws.iter_rows(values_only=True):
    print(row)


# update the data in the worksheet
ws['B2'] = 26
workbook.save('example.csv')
